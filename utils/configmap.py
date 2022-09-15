import os
import re
import pandas as pd
from openpyxl import load_workbook
from ruamel.yaml import ruamel

tempDict = dict()
tempDict["kind"] = "ConfigMap"
tempDict["metadata"] = {"name": "dut-a-cfg"}
tempDict["data"] = {"mg.cfg": "|"}


def createConfigMap(line):
    result = re.findall(r'\{.*?}\}', line)
    filteredResult = []
    for index in range(len(result)):
        key = result[index]
        filteredResult.append(key)
    return filteredResult


def readAllTemplates(dut_config_path):
    pwd = os.getcwd() + "/pactemplates/"
    arr = os.listdir(pwd)
    provisionedTemplates = os.getcwd() + "/Input/PAC_LLD.xlsx"
    finalSheetList = []
    workbook = load_workbook(provisionedTemplates, read_only=True)
    sheetName = "Provisioning"
    if 'Provisioning_PAC' in workbook.sheetnames:
        sheetName = "Provisioning_PAC"
    sheet_to_df_map = pd.read_excel(provisionedTemplates, sheet_name=sheetName)
    sheet_list = sheet_to_df_map['Unnamed: 1'].tolist()[2:]
    provisioningValue = sheet_to_df_map['Unnamed: 2'].tolist()[2:]
    for sheetIndex in range(len(sheet_list)):
        if provisioningValue[sheetIndex] == "CREATE":
            finalSheetList.append(sheet_list[sheetIndex])
    exceptionList = []
    configMap = ""
    for item in range(len(finalSheetList)):
        templateName = finalSheetList[item]
        finalTemplateName = templateName.replace("-", "_")
        content = "\n    #------------------------------------\n    echo " + templateName + " Configuration\n    " \
                                                                                            "#------------------------------------\n "
        content = content + "\n{{- if .Values." + finalTemplateName + " }}" \
                                                                      "\n{{- range .Values." + finalTemplateName + " }}"
        try:
            with open(pwd + templateName + "_template") as file:
                lines = file.readlines()
                if not lines[0].startswith("/configure"):
                    continue
                stringContent = ""
                endString = "{{- end}}"
                for index in range(len(lines)):
                    configMapContent = ""
                    result = createConfigMap(lines[index])
                    value = lines[index].strip()
                    if len(result) > 0:
                        for keysIndex in range(len(result)):
                            if result[keysIndex].__contains__("-") or result[keysIndex].__contains__(" "):
                                resultKeysIndex = result[keysIndex].replace("-", "_")
                                resultKeysIndex = resultKeysIndex.replace(" ", "_")
                                if resultKeysIndex.__contains__("(") or resultKeysIndex.__contains__("["):
                                    resultKeysIndex = re.sub("\(.*?\)", "()", resultKeysIndex)
                                    resultKeysIndex = resultKeysIndex.replace("()", "")
                                    resultKeysIndex = re.sub("\(.*?\)", "[]", resultKeysIndex)
                                    resultKeysIndex = resultKeysIndex.replace("[]", "")
                            else:
                                resultKeysIndex = result[keysIndex]
                            resultKeysIndex = resultKeysIndex.replace("{{per['", "")
                            resultKeysIndex = resultKeysIndex.replace("']}}", "")
                            resultKeysIndex = resultKeysIndex.replace("{{per[\"", "")
                            resultKeysIndex = resultKeysIndex.replace("\"]}}", "")
                            resultKeysIndex = "".join(("parameter_", resultKeysIndex)) if resultKeysIndex[
                                                                                          :1].isdigit() else resultKeysIndex
                            replaceWith = "{{." + resultKeysIndex + "}}"
                            value = value.replace(result[keysIndex], replaceWith)
                            if len(configMapContent) == 0:
                                configMapContent = "." + resultKeysIndex
                            else:
                                configMapContent = configMapContent + " " + "." + resultKeysIndex
                        if len(result) > 1:
                            configMapContent = "{{- if and " + configMapContent + "}}"
                        else:
                            configMapContent = "{{- if " + configMapContent + "}}"
                    if len(stringContent) != 0:
                        if len(result) > 0:
                            stringContent = stringContent + "\n" + configMapContent + "\n    " + value + "\n" + endString
                        else:
                            stringContent = stringContent + "\n" + configMapContent + "    " + value
                    else:
                        if len(result) > 0:
                            stringContent = configMapContent + "\n    " + value + "\n" + endString
                        else:
                            stringContent = configMapContent + "    " + value
                stringContent = content + "\n" + stringContent + "\n" + endString + "\n" + endString
                configMap = configMap + "\n" + stringContent
        except:
            exceptionList.append(arr[item])
            continue
    # print(configMap)
    if not os.path.exists(dut_config_path):
        os.remove(dut_config_path)
    with open(dut_config_path, "w") as yamlFile:
        ruamel.yaml.round_trip_dump(tempDict, yamlFile, default_flow_style=False)
    with open(dut_config_path, "a") as yaml_file:
        yaml_file.write(configMap)
    # print("Exception List: \n", exceptionList)
