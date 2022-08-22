import os
from itertools import islice

from openpyxl import load_workbook
from ruamel.yaml import ruamel
from ruamel.yaml.comments import CommentedMap as OrderedDict
from unflatten import unflatten

cwd = os.getcwd()


def merge(dict1, dict2):
    dict3 = {**dict1, **dict2}
    for key, value in dict3.items():
        if key in dict1 and key in dict2:
            dict3[key] = dict1[key]
            if not isinstance(dict3[key], list):
                dict3[key] = [dict3[key]]
            dict3[key].append(value)
    return dict3


def process(ciqsheetpath):
    wb = load_workbook(ciqsheetpath)
    sheetnames = wb.get_sheet_names()
    prov_sheet = wb["Provisioning_TBT"]
    create_sheets = []
    alias_sheets = {}

    for rowitems in islice(prov_sheet.values, 2, prov_sheet.max_row):
        if rowitems[2] is not None and rowitems[2] == 'CREATE':
            create_sheets.append(rowitems[1])
            try:
                if rowitems[3] == None:
                    alias_sheets[rowitems[1]] = rowitems[1]
                else:
                    alias_sheets[rowitems[1]] = rowitems[3]
            except:
                alias_sheets[rowitems[1]] = rowitems[1]

    eligible_sheet_names = []
    if len(create_sheets) > 0:
        for sheetnametemp in sheetnames:
            if sheetnametemp in create_sheets:
                sheetname = sheetnametemp
                eligible_sheet_names.append(sheetname)
                alias = alias_sheets[sheetname]
                values_yaml = cwd + "/output/"+alias+".yaml"
                if not os.path.exists(values_yaml):
                    open(values_yaml, 'w').close()
                sheet = wb[sheetname]

                excel_list = []
                excel_dict = {}

                for row in islice(sheet.values, 1, sheet.max_row):
                    excel = OrderedDict()
                    if row[1] is not None:
                        str = row[0]
                        excel[str] = row[1]
                        excel_list.append(excel)

                for i in excel_list:
                    excel_dict = merge(excel_dict, i)

                excel_dict = unflatten(excel_dict)

                with open(values_yaml, "w") as yamlFile:
                    ruamel.yaml.round_trip_dump(excel_dict, yamlFile, default_flow_style=False)
    return eligible_sheet_names, alias_sheets